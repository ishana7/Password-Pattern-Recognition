#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Sep 22 22:03:40 2018

@author: ishana798
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Sep 22 20:35:46 2018

@author: ishana798
"""

# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import openpyxl
from tkinter import *
from time import time
import os
from PIL import Image,ImageDraw,ImageTk
creds = 'tempfile.temp'

global j
j=0
global y
y=1
def wel():
    global roots
    print("hello")
    roots = Tk()
    w=800
    h=600
    width,height=roots.winfo_screenwidth(),roots.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    roots.geometry('%dx%d+%d+%d'%(w,h,x,y))
    roots.configure(bg="#90C3C8")
    roots.title('WELCOME')
    intruction = Label(roots,font=("Poor Richard",40), text='                                   WELCOME\n',highlightthickness=5,highlightbackground="black",bg="#90C3C8")
    intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)

    loginB = Button(roots, text='Login', command=Login,fg="#EA7317",bg="#EA7317",height=3,width=20)
    #loginB.grid(row=4, column=0)
    loginB.place(relx=.41,rely=.3)
    signupB = Button(roots, text='Signup', command=Signup,fg="#EA7317",bg="#EA7317",height=3,width=20)
    #signupButton.grid(row=1, column=1)
    signupB.place(relx=.41,rely=.5)
    analysisB = Button(roots, text='Analysis', command=analysis,fg="#EA7317",bg="#EA7317",height=3,width=20)
    analysisB.place(relx=.41,rely=.7)
    roots.mainloop()

def Login():
    global pwordE
    global nameE
    global roots

    roots.destroy()
    roots = Tk()
    w=800
    h=600
    width,height=roots.winfo_screenwidth(),roots.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    roots.geometry('%dx%d+%d+%d'%(w,h,x,y))
    roots.configure(bg="#A0CCDA")
    roots.title('LOGIN')
    intruction = Label(roots,font=("Poor Richard",40), text='                                   Login\n',highlightthickness=5,highlightbackground="black",bg="#A0CCDA")
    intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)


    nameL = Label(roots,font=("",20), text='Username   :: ',highlightthickness=5,bg="#A0CCDA")
    nameL.place(relx=.2,rely=.25)

    nameE = Entry(roots,bg="lightgrey",highlightthickness=2,highlightbackground="grey")
    nameE.place(relx=0.5,rely=0.28)

    loginB = Button(roots, text='confirm', command=gett,fg="#EA7317",bg="#EA7317",height=3,width=20)
    loginB.place(relx=.42,rely=.6)
    roots.mainloop()
def gett():
    global rot
    global roots
    global counter
    global chk
    global cmp
    global totaltime
    rot=Tk()
    w=800
    h=600
    width,height=rot.winfo_screenwidth(),rot.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    rot.geometry('%dx%d+%d+%d'%(w,h,x,y))
    rot.configure(bg="#A0CCDA")

    mylist = []
    file = 'data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb["Sheet1"]
    rows=ws.max_row
    for i in range(1,rows+1):
        mylist.extend([ws.cell(row=i,column=1).value])

    pwordL = Label(rot,font=("",20), text='Password   :: ',highlightthickness=5,bg="#A0CCDA")
    pwordL.place(relx=.2,rely=.25)
    pwordE = Entry(rot,bg="lightgrey",highlightthickness=2,highlightbackground="grey",show='*')
    pwordE.place(relx=0.5,rely=0.28)
    tw=pwordE.get()
    name=nameE.get()

    if name in mylist:
        print="bye"
        counter=mylist.index(name)
        counter+=1
        cmp=ws.cell(row=counter,column=2).value

        roots.destroy()
        #ins = Label(rot, text='username confirmed ')
        #ins.grid(row=1, column=0, sticky=W)

        loginB = Button(rot, text='Login', command=CheckLogin,fg="#EA7317",bg="#EA7317",height=3,width=20)
        loginB.place(relx=.42,rely=.6)
        def logg(keyevent):
            global totaltime
            cword = cmp
            cwordsize = len(cword)
            cwordlist = tuple(cword)
            cwordfl = str(cwordlist[0])
            cwordll = str(cwordlist[-1])

            tword = pwordE.get()
            twordsize = len(tword)
            if twordsize > 0:
                twordlist = tuple(tword)
            twordfl = str(twordlist[0])
            twordll = str(twordlist[-1])
            if cwordsize == 1 and twordsize == 1:
                print("more letters")
            if twordsize == 1 and cwordsize > 1:
                global start
                start = time()
            if twordsize == cwordsize and twordsize != 1:
                if cwordll == twordll:
                    stop = time()
                    totaltime = stop - start
            new_col=[tword]
            file='data.xlsx'
            wb = openpyxl.load_workbook(filename=file)
            ws = wb["Sheet1"]
            col=11
            for row, entry in enumerate(new_col, start=1):
                ws.cell(row=counter, column=col, value=entry)

            wb.save(file)
            wb.close()

        pwordE.bind('<KeyRelease>', logg)
        #print(tword)

def CheckLogin():
    ##############################
    global cmp
    global totaltime
    rot.destroy()
    file = 'data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb["Sheet1"]
    tw=ws.cell(row=counter,column=11).value
    t1=ws.cell(row=counter,column=9).value
    t2=ws.cell(row=counter,column=10).value

    if tw == cmp and t1<=totaltime and totaltime<=t2:
        r = Tk()
        r.title(':D')
        r.geometry('150x150')
        rlbl = Label(r, text='\n[+] Logged In')
        rlbl.pack()
        r.mainloop()
    else:
        r = Tk()
        r.title('D:')
        r.geometry('150x150')
        rlbl = Label(r, text='\n[!] Invalid Login')
        rlbl.pack()
        r.mainloop()




def Signup():
    global nameEL
    global pwordEL
    global emailEL
    global mobEL
    global ageEL
    global rootA
    global roots
    global rows

    roots.destroy()
    fields = 'Username', 'Password', 'Email', 'Mob','Age'
    def fetch(entries):
        file = 'data.xlsx'
        wb = openpyxl.load_workbook(filename=file)
        ws = wb["Sheet1"]
        rows=ws.max_row
        rows+=1

        col=0

        for entry in entries:
            text  = entry[1].get()

            new_col = [text]

            col+=1

            for row, entry in enumerate(new_col, start=1):
                ws.cell(row=rows, column=col, value=entry)

            wb.save(file)

    def makeform(root, fields):
       entries = []
       for field in fields:
          row = Frame(rootA)
          lab = Label(row, width=35,font=("",13), text=field+"  ::", anchor='w',bg="#90C3C8")
          ent = Entry(row)
          row.pack(side=TOP, fill=Y, padx=5, pady=5)
          lab.pack(side=LEFT)
          ent.pack(side=RIGHT, expand=YES, fill=Y)
          entries.append((field, ent))
       return entries

    if __name__ == '__main__':
       fields = 'Username', 'Password', 'Email', 'Mob','Age'
       rootA = Tk()
       w=800
       h=600
       width,height=rootA.winfo_screenwidth(),rootA.winfo_screenheight()
       x=(width/2)-(w/2)
       y=(height/2)-(h/2)
       rootA.geometry('%dx%d+%d+%d'%(w,h,x,y))
       rootA.configure(bg="#90C3C8")

       rootA.title('SIGN-UP')
       intruction = Label(rootA,font=("",20), text='Please Enter  Credentials\n',bg="#90C3C8")
       intruction.pack(side=LEFT, padx=5, pady=5)
       ents = makeform(rootA, fields)
       rootA.bind('<Return>', (lambda event, e=ents: fetch(e)))

       #b1 = Button(rootA, text='Save',command=(lambda e=ents: fetch(e)),fg="#a1dbcd",bg="#383a39",height=3,width=10)
       #b1.place(relx=.32,rely=.8)

       b2 = Button(rootA, text='Next',command=combine_funcs((lambda e=ents: fetch(e)), sign),fg="#EA7317",bg="#EA7317",height=3,width=10)
       b2.place(relx=0.42,rely=0.8)

       rootA.mainloop()

def combine_funcs(*funcs):
    def combined_func(*args, **kwargs):
        for f in funcs:
            f(*args, **kwargs)
    return combined_func

def sign():
     rootA.destroy()
     pas()

def DelUser():
    os.remove(creds)
    rootA.destroy()
    Login()
def paes():
    global master
    master.destroy()
    pas()
def exi():
    global master
    master.destroy()
    Login()

def pas():
    global j
    j+=1
    global master

    master = Tk()
    w=700
    h=500
    width,height=master.winfo_screenwidth(),master.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    master.geometry('%dx%d+%d+%d'%(w,h,x,y))
    master.configure(bg="#A0CCDA")
    master.title('LOGIN')
    
    intruction = Label(master,font=("Poor Richard",30), text='       CONFIRM     PASSWORD \n',highlightthickness=5,highlightbackground="black",bg="#A0CCDA")
    intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)
    
    w1 = Label(master,font=("",20), text='Password   :: ',highlightthickness=5,bg="#A0CCDA")    
    w1.grid(row=3, column=0,padx=10,pady=2, sticky=N+S)   
    
    we = Entry(master,bg="lightgrey",highlightthickness=2,highlightbackground="grey")
    we.grid(row=3, column=1,padx=10,pady=2, sticky=N+S)
    
   

    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb['Sheet1']
    for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
        for cell in row:
            continue
        ce=cell.value
    sheet = wb.active

    def pressed(keyevent):
        cword = ce
        cwordsize = len(cword)
        cwordlist = tuple(cword)
        cwordfl = str(cwordlist[0])
        cwordll = str(cwordlist[-1])

        tword = we.get()
        twordsize = len(tword)
        if twordsize > 0:
            twordlist = tuple(tword)
        twordfl = str(twordlist[0])
        twordll = str(twordlist[-1])
        if cwordsize == 1 and twordsize == 1:
            print("more letters")
        if twordsize == 1 and cwordsize > 1:
            global start
            start = time()
        if twordsize == cwordsize and twordsize != 1:
            if cwordll == twordll:
                stop = time()
                totaltime = stop - start

                file = 'data.xlsx'
                new_col = [totaltime]

                wb = openpyxl.load_workbook(filename=file)
                ws = wb["Sheet1"]
                rows = ws.max_row
                if j==1:
                    col = 6
                if j==2:
                    col = 7
                if j==3:
                    col = 8


                for row, entry in enumerate(new_col, start=1):
                    ws.cell(row=rows, column=col, value=entry)

                wb.save(file)
                wb.close()



    we.bind('<KeyRelease>', pressed)
    
    if j<3:
        okButton = Button(master, text='confirm', command=pas,fg="#EA7317",bg="#EA7317",height=3,width=20)
        okButton.place(relx=.42,rely=.6)

    else:
       ekButton = Button(master, text='end', command=cal,fg="#EA7317",bg="#EA7317",height=3,width=20)
       ekButton.place(relx=.42,rely=.6)

def cal():
    global master

    wb = openpyxl.load_workbook('data.xlsx')
    ws = wb['Sheet1']
    rows=ws.max_row
    sum=0
    sheet = wb.active
    for col in ws.iter_cols(min_row=rows,max_row=rows,min_col=6,max_col=8):
        for cell in col:
            sum=sum+cell.value

    avg=sum/3
    minavg=avg-0.1
    max_avg=avg+0.1

    new_col = [minavg]
    file='data.xlsx'
    wb = openpyxl.load_workbook(filename=file)
    ws = wb["Sheet1"]
    col=9

    for row, entry in enumerate(new_col, start=1):
        ws.cell(row=rows, column=col, value=entry)
    new_col = [max_avg]
    col=10
    for row, entry in enumerate(new_col, start=1):
        ws.cell(row=rows, column=col, value=entry)

    wb.save(file)
    wb.close()
    master.destroy()
    Login()

def analysis():
    print('helo')
    global master1
    
    
    master1 = Tk()
    w=700
    h=500
    width,height=master1.winfo_screenwidth(),master1.winfo_screenheight()
    x=(width/2)-(w/2)
    y=(height/2)-(h/2)
    master1.geometry('%dx%d+%d+%d'%(w,h,x,y))
    #master1.configure(bg="#A0CCDA")
    #master1.title('ANALYSIS')
    
    # = Label(master1,font=("Poor Richard",30), text='                    ANALYSIS\n',highlightthickness=5,highlightbackground="black",bg="#A0CCDA")
    #intruction.grid(row=0, column=0,padx=10,pady=2, sticky=N+S)
    
    #canvas = Canvas(master1, width = 200, height = 200)      
    #canvas.pack()      
    img = ImageTk.PhotoImage(Image.open("analy1.png"))   
    im = Label(master1,image=img,compound=BOTTOM)
    im.pack()
    master1.mainloop()

if os.path.isfile(creds):
    wel()
else:
    wel()
